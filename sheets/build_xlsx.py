"""Build jwal-beer-review.xlsx from the CSVs in this directory.

Produces a single workbook with four tabs:
  1. Instructions  (green tab, tab #1) - where to put new rows
  2. Beers         (green tab) - PUBLISHES to the live site
  3. Breweries     (grey tab)  - reference
  4. Locations     (grey tab)  - reference

Run: python3 sheets/build_xlsx.py
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
OUT = HERE / "jwal-beer-review.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2E7D32")          # dark green
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
REF_HEADER_FILL = PatternFill("solid", fgColor="455A64")      # slate
NOTE_FILL = PatternFill("solid", fgColor="FFF59D")            # pale yellow
TITLE_FONT = Font(bold=True, size=16, color="1B5E20")
BOLD = Font(bold=True)

def load_csv(name):
    with (HERE / name).open(newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    return rows[0], rows[1:]

def style_header(ws, ncols, fill=HEADER_FILL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

def autosize(ws, headers, rows):
    for i, h in enumerate(headers, 1):
        max_len = len(str(h))
        for r in rows:
            if i - 1 < len(r):
                max_len = max(max_len, len(str(r[i - 1])))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)

wb = Workbook()

# ---- Instructions tab ----
ws = wb.active
ws.title = "Instructions"
ws.sheet_properties.tabColor = "2E7D32"

ws["A1"] = "JWAL Beer Review - Data Entry"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:B1")

lines = [
    ("", ""),
    ("WHERE TO ADD NEW REVIEWS", ""),
    ("Open the 'Beers' tab (green tab below).", ""),
    ("Scroll to the first empty row and type your review.", ""),
    ("That's it - the live site re-fetches on every page load.", ""),
    ("", ""),
    ("ONLY the Beers tab publishes to the site.", ""),
    ("Breweries / Locations are lookup references - edit only if you", ""),
    ("add a brewery or city that isn't already listed.", ""),
    ("", ""),
    ("REQUIRED COLUMNS (Beers tab)", ""),
    ("beer", "Beer name as marketed"),
    ("style", "Lager / Pilsner / Wheat Beer / Belgian Ale / IPA / Pale Ale / Stout / Brown Ale / Red Ale"),
    ("origin", "ISO-2 country code of the brewery (e.g. NL, DE, US)"),
    ("abv", "Alcohol %, decimal (e.g. 5.0)"),
    ("method", "Bottle / Can / Draft / Nitro"),
    ("city", "City where consumed"),
    ("region", "Region / state where consumed"),
    ("country", "Country where consumed (full name, e.g. USA)"),
    ("cc", "ISO-2 country code where consumed (e.g. US)"),
    ("rating", "0.00 - 5.00 in 0.25 increments"),
    ("isNew", "TRUE if first time drinking this beer, otherwise FALSE"),
    ("month", "3-letter month: Jan, Feb, Mar..."),
    ("monthN", "Month number 1-12 (matches 'month')"),
    ("year", "4-digit year"),
    ("", ""),
    ("PUBLISHING", ""),
    ("File -> Share -> Publish to web -> Entire document, CSV.", ""),
    ("The site is already wired to this sheet's ID.", ""),
]

for i, (a, b) in enumerate(lines, start=2):
    ws.cell(row=i, column=1, value=a)
    ws.cell(row=i, column=2, value=b)

# Bold the section headers and key callout
for row in (3, 7, 12, 28):
    ws.cell(row=row, column=1).font = BOLD

# Highlight the "ONLY the Beers tab publishes" note
for row in (7, 8, 9, 10):
    for col in (1, 2):
        ws.cell(row=row, column=col).fill = NOTE_FILL

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 80

# ---- Beers tab ----
headers, rows = load_csv("beers.csv")
ws = wb.create_sheet("Beers")
ws.sheet_properties.tabColor = "2E7D32"
ws.append(headers)
for r in rows:
    ws.append(r)

# Comment on A1 reminding user where to add
ws["A1"].comment = Comment(
    "ADD NEW REVIEWS in the first empty row below.\n"
    "This tab is what publishes to the live site.",
    "JWAL"
)

style_header(ws, len(headers))
autosize(ws, headers, rows)

# Convert types: numeric columns
numeric_cols = {"abv": 4, "rating": 10, "monthN": 13, "year": 14}
bool_col = 11  # isNew
for row_idx in range(2, len(rows) + 2):
    for name, col in numeric_cols.items():
        cell = ws.cell(row=row_idx, column=col)
        try:
            cell.value = float(cell.value) if "." in str(cell.value) else int(cell.value)
        except (TypeError, ValueError):
            pass
    b = ws.cell(row=row_idx, column=bool_col)
    if str(b.value).lower() == "true":
        b.value = True
    elif str(b.value).lower() == "false":
        b.value = False

# Data validation dropdowns
def add_dv(ws, col_letter, formula, first_row=2, last_row=2000):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")
    ws.add_data_validation(dv)

add_dv(ws, "B", '"Lager,Pilsner,Wheat Beer,Belgian Ale,IPA,Pale Ale,Stout,Brown Ale,Red Ale"')
add_dv(ws, "E", '"Bottle,Can,Draft,Nitro"')
add_dv(ws, "K", '"TRUE,FALSE"')
add_dv(ws, "L", '"Jan,Feb,Mar,Apr,May,Jun,Jul,Aug,Sep,Oct,Nov,Dec"')

# ---- Breweries tab ----
headers, rows = load_csv("breweries.csv")
ws = wb.create_sheet("Breweries")
ws.sheet_properties.tabColor = "607D8B"
ws.append(headers)
for r in rows:
    ws.append(r)
style_header(ws, len(headers), fill=REF_HEADER_FILL)
autosize(ws, headers, rows)
for row_idx in range(2, len(rows) + 2):
    for col in (7, 8):  # lat, lng
        c = ws.cell(row=row_idx, column=col)
        try:
            c.value = float(c.value)
        except (TypeError, ValueError):
            pass

# ---- Locations tab ----
headers, rows = load_csv("locations.csv")
ws = wb.create_sheet("Locations")
ws.sheet_properties.tabColor = "607D8B"
ws.append(headers)
for r in rows:
    ws.append(r)
style_header(ws, len(headers), fill=REF_HEADER_FILL)
autosize(ws, headers, rows)
for row_idx in range(2, len(rows) + 2):
    for col in (5, 6):  # lat, lng
        c = ws.cell(row=row_idx, column=col)
        try:
            c.value = float(c.value)
        except (TypeError, ValueError):
            pass

wb.save(OUT)
print(f"wrote {OUT}")
